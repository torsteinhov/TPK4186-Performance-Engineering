import os
from openpyxl import load_workbook
from datetime import datetime

"""
    A python file with managing functions for excel spreadsheets. Adds multiple functionalities like returning all files in folder, extracts excel data and calculations based on
    datetime objects.

    Methods
    -------
    listFilesInFolder()
        returns all files in a given folder.
    extractFiles()
        loads an excel workbook and iterates through it by using a triple for loop to access sheets, rows and cells.
        Enumerate helps us keep track of the rows and the function returns a dictionary with the format {row: [cell1, cell2, cell3, ...]}.
    dateFromString(string)
        returns a datetime object based on the string input.
    diffDate(date1, date2)
        takes in two datetime objects, calculates the time difference between the dates and returns it in hours format.
"""

def listFilesInFolder(folder):

    files = os.listdir(folder)

    files_list = []

    for excel_file in files:

        files_list.append(str(excel_file))

    return files_list


def extractExcel(workbook):

    workbook_dict = {}

    workbook = load_workbook(workbook)
    for worksheet in workbook:
        for index, row in enumerate(worksheet.iter_rows()):
            content = [] 
            for cell in row:
                content.append(cell.value)
            workbook_dict[index+1] = content
    
    return workbook_dict


def dateFromString(string):
    '''
    Not really necessary since our extractExcel() function parses and returns the relevant cells
    in datetime format.
    '''

    date_time_object = datetime.strptime(string, '%Y-%m-%d  %H:%M:%S')

    return date_time_object

def diffDate(date1, date2):
    '''
    date1 : datetime_object
    date2 : datetime_object
    '''

    duration = date2 - date1
    duration_seconds = duration.total_seconds()
    hours = divmod(duration_seconds, 3600)[0]

    return hours


def testFunctions():

    listFilesInFolder('ReliabilityData')
    print(extractExcel('ReliabilityData/Plant1.xlsx'))
    print(dateFromString('14-02-2021  14:00:00'))

    date1 = dateFromString('14-02-2021  14:00:00')
    date2 = dateFromString('16-02-2021  14:00:00')
    print(diffDate(date1, date2))
    

testFunctions()